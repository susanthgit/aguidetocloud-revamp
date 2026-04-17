"""Generate Ko-fi product update spreadsheet with optimised titles, descriptions, and cross-links."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ko-fi Product Updates"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1a1a2e")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

FOOTER = """

━━━━━━━━━━━━━━━━━━━━━━━
MORE FREE RESOURCES
━━━━━━━━━━━━━━━━━━━━━━━
28+ Free Cloud & AI Tools: aguidetocloud.com/links/
AI News (updated 4x daily): aguidetocloud.com/ai-news/
52 Free Cert Study Guides: aguidetocloud.com/cert-tracker/
YouTube Full Courses: youtube.com/@SusanthSutheesh
YouTube Quick Bites: youtube.com/@AGuideToCloud
Blog: aguidetocloud.com/blog/

Made with love by A Guide to Cloud & AI
aguidetocloud.com"""

headers = ["#", "Thumbnail File", "Current Title", "NEW Title", "NEW Description", "Price", "Category"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 80
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 14

items = [
    # (thumbnail, old_title, new_title, description, price, category)
    # --- EXAM Q&A ---
    ("ai900-qa.jpg", "AI-900 Exam 100 Q&As with explanation",
     "AI-900 Exam Q&A — 100 Questions with Explanations",
     "Preparing for the AI-900: Microsoft Azure AI Fundamentals exam?\n\nThis pack includes 100 practice questions with detailed explanations covering all exam objectives.\n\nWhat you get:\n- 100 questions with full explanations\n- PDF, Word & PowerPoint flash card formats\n- Covers: AI workloads, ML principles, Azure AI services, Responsible AI\n\nPerfect for self-study, revision, or mock exam practice.",
     "$5+", "Exam Q&A"),

    ("sc100-qa.jpg", "SC-100 Exam 100 Q&As with explanation",
     "SC-100 Exam Q&A — 100 Questions with Explanations",
     "Preparing for the SC-100: Microsoft Cybersecurity Architect exam?\n\nThis pack includes 100 practice questions with detailed explanations.\n\nWhat you get:\n- 100 expert-level questions with explanations\n- PDF, Word & PowerPoint flash card formats\n- Covers: Zero Trust, identity security, security operations, data security\n\nDesigned for experienced security professionals preparing for this expert-level exam.",
     "$5+", "Exam Q&A"),

    ("pl900-qa.jpg", "PL-900 Exam Q&A (38 Questions)",
     "PL-900 Exam Q&A — 38 Practice Questions",
     "Practice questions for the PL-900: Microsoft Power Platform Fundamentals exam.\n\nWhat you get:\n- 38 questions covering all exam objectives\n- Covers: Power Apps, Power Automate, Power BI, Power Virtual Agents, Dataverse\n\nGreat for a quick revision before your exam.",
     "$5+", "Exam Q&A"),

    ("ms900-qa.jpg", "MS-900 Exam 127 Q&As with explanation",
     "MS-900 Exam Q&A — 127 Questions with Explanations",
     "The most comprehensive MS-900 practice pack available!\n\nWhat you get:\n- 127 questions (100 NEW + 27 bonus from previous batch)\n- Detailed explanations for every answer\n- PDF, Word & PowerPoint flash card formats\n- Covers: Cloud concepts, Microsoft 365 apps, security, compliance, licensing\n\nOur best-selling M365 fundamentals resource.",
     "$5+", "Exam Q&A"),

    ("ms500-qa.jpg", "MS-500 Exam Q&A (89 Questions)",
     "MS-500 Exam Q&A — 89 Practice Questions (FREE)",
     "FREE download — 89 practice questions for the MS-500: Microsoft 365 Security Administration exam.\n\nWhat you get:\n- 89 questions covering all exam domains\n- Covers: Identity & access, threat protection, information protection, governance\n\nFree resource to help you prepare. Pay what you want if you find it valuable!",
     "Free+", "Exam Q&A"),

    ("dp900-qa.jpg", "DP-900 Exam Q&A (39 Questions)",
     "DP-900 Exam Q&A — 39 Practice Questions",
     "Practice questions for the DP-900: Microsoft Azure Data Fundamentals exam.\n\nWhat you get:\n- 39 questions covering all exam objectives\n- Covers: Core data concepts, relational data, non-relational data, analytics workloads\n\nQuick and focused revision material.",
     "$5+", "Exam Q&A"),

    ("az900-qa.jpg", "AZ-900 Exam Q&A (53 Questions)",
     "AZ-900 Exam Q&A — 53 Practice Questions",
     "Practice questions for the AZ-900: Microsoft Azure Fundamentals exam.\n\nWhat you get:\n- 53 questions covering all exam objectives\n- Covers: Cloud concepts, Azure architecture, management & governance\n\nPair with our free AZ-900 study guide at aguidetocloud.com/cert-tracker/az-900/",
     "$5+", "Exam Q&A"),

    ("az500-qa.jpg", "AZ-500 Exam Q&A (73 Questions)",
     "AZ-500 Exam Q&A — 73 Practice Questions",
     "Practice questions for the AZ-500: Microsoft Azure Security Technologies exam.\n\nWhat you get:\n- 73 questions covering all exam objectives\n- Covers: Identity management, platform protection, security operations, data protection\n\nGreat companion to our AZ-500 YouTube full course.",
     "$5+", "Exam Q&A"),

    ("az400-qa.jpg", "AZ-400 Exam Q&A (53 Questions)",
     "AZ-400 Exam Q&A — 53 Practice Questions",
     "Practice questions for the AZ-400: Microsoft DevOps Solutions exam.\n\nWhat you get:\n- 53 questions covering all exam objectives\n- Covers: Source control, CI/CD, dependency management, infrastructure as code, monitoring\n\nGreat companion to our AZ-400 YouTube full course.",
     "$5+", "Exam Q&A"),

    ("az305-qa.jpg", "AZ-305 Exam Q&A (99 Questions)",
     "AZ-305 Exam Q&A — 99 Practice Questions",
     "Practice questions for the AZ-305: Microsoft Azure Infrastructure Solutions exam.\n\nWhat you get:\n- 99 questions covering all exam objectives\n- Covers: Governance, compute, storage, networking, business continuity, data integration\n\nNearly 100 questions — our most comprehensive Azure architect prep pack.",
     "$5+", "Exam Q&A"),

    ("az204-qa.jpg", "AZ-204 Exam Q&A (49 Questions)",
     "AZ-204 Exam Q&A — 49 Practice Questions",
     "Practice questions for the AZ-204: Developing Solutions for Microsoft Azure exam.\n\nWhat you get:\n- 49 questions covering all exam objectives\n- Covers: Azure compute, storage, security, monitoring, third-party integrations\n\nFocused developer-track preparation.",
     "$5+", "Exam Q&A"),

    ("az104-qa.jpg", "AZ-104 Exam Q&A (92 Questions)",
     "AZ-104 Exam Q&A — 92 Practice Questions",
     "Practice questions for the AZ-104: Microsoft Azure Administrator exam.\n\nWhat you get:\n- 92 questions covering all exam objectives\n- Covers: Identity, governance, storage, compute, virtual networking, monitoring\n\nOur most popular Azure admin resource. Pair with our free study guide at aguidetocloud.com/cert-tracker/az-104/",
     "$5+", "Exam Q&A"),

    ("sc900-qa.jpg", "SC-900 Exam Q&A (50 Questions)",
     "SC-900 Exam Q&A — 50 Practice Questions",
     "Practice questions for the SC-900: Microsoft Security, Compliance & Identity Fundamentals exam.\n\nWhat you get:\n- 50 questions covering all exam objectives\n- Covers: Security, compliance, identity concepts, Microsoft Entra, Defender, Purview\n\nPerfect entry-point for security certifications.",
     "$5+", "Exam Q&A"),

    # --- INTERVIEW Q&A ---
    ("azure-interview.jpg", "100 Azure Interview Q&As",
     "100 Azure Interview Q&As — Complete Guide",
     "Getting ready for an Azure job interview?\n\nThis 138+ page guide covers 100 real-world interview questions with detailed explanations.\n\nWhat you get:\n- 100 Q&As organised by topic\n- PDF (138+ pages), Word & PowerPoint flash cards\n- Topics: Cloud concepts, VMs, networking, storage, security, identity, monitoring, cost management\n\nUsed by 66+ people to prepare for Azure roles.",
     "$5+", "Interview"),

    ("m365-interview.jpg", "100 Microsoft 365 Interview Q&As",
     "100 Microsoft 365 Interview Q&As — Complete Guide",
     "Preparing for a Microsoft 365 job interview?\n\nThis 140+ page guide covers 100 real-world interview questions with detailed explanations.\n\nWhat you get:\n- 100 Q&As organised by topic\n- PDF (140+ pages), Word & PowerPoint flash cards\n- Topics: Exchange Online, SharePoint, Teams, Security, Compliance, Intune, Entra ID\n\nCovers both admin and end-user perspectives.",
     "$5+", "Interview"),

    ("avd-w365-interview.jpg", "AVD vs W365 Interview Q&As",
     "AVD vs Windows 365 Interview Q&As — Comparison Guide",
     "Specialising in virtual desktop technologies?\n\nThis guide covers interview questions comparing Azure Virtual Desktop (AVD) and Windows 365.\n\nWhat you get:\n- Detailed comparison Q&As with explanations\n- PDF, Word & PowerPoint flash card formats\n- Covers: Architecture, licensing, management, use cases, cost comparison\n\nEssential for EUC/VDI specialist interviews.",
     "$5+", "Interview"),

    # --- BOOTCAMP LABS ---
    ("az700-bootcamp.jpg", "Azure Networking [AZ-700] Bootcamp",
     "AZ-700 Bootcamp — Azure Networking Labs (11 Labs, 116 Exercises)",
     "Hands-on lab guide for the AZ-700: Designing and Implementing Azure Networking Solutions exam.\n\nWhat you get:\n- 11 structured labs with 116 exercises\n- Step-by-step instructions with screenshots\n- Covers: VNets, peering, VPN gateways, ExpressRoute, Load Balancers, Azure Firewall, DNS, Traffic Manager\n\nOur #1 best-selling bootcamp — used by 148+ people.",
     "$5+", "Bootcamp"),

    ("az140-bootcamp.jpg", "Azure Virtual Desktop [AZ-140] Bootcamp",
     "AZ-140 Bootcamp — Azure Virtual Desktop Labs (12 Labs, 178 pages)",
     "Hands-on lab guide for the AZ-140: Configuring and Operating Azure Virtual Desktop exam.\n\nWhat you get:\n- 12 structured labs with 26 exercises\n- 178 pages of step-by-step instructions\n- Covers: Host pools, session hosts, user profiles, FSLogix, security, monitoring, scaling\n\nThe most complete AVD lab guide available.",
     "$5+", "Bootcamp"),

    ("az900-bootcamp.jpg", "Azure Fundamental [AZ-900] Bootcamp",
     "AZ-900 Bootcamp — Azure Fundamentals Labs (21 Labs, 89 pages)",
     "Beginner-friendly hands-on lab guide for the AZ-900: Microsoft Azure Fundamentals exam.\n\nWhat you get:\n- 21 labs with 56 exercises\n- 89 pages of step-by-step instructions\n- Covers: Resource groups, VMs, storage, networking, databases, App Services, Azure AD\n\nPerfect first lab experience for Azure beginners.",
     "$5+", "Bootcamp"),

    ("sentinel-bootcamp.jpg", "Microsoft Sentinel Bootcamp",
     "Microsoft Sentinel Bootcamp — SIEM Labs (8 Labs, 104 pages)",
     "Hands-on lab guide for Microsoft Sentinel — Azure's cloud-native SIEM.\n\nWhat you get:\n- 8 structured labs with 29 exercises\n- 104 pages of step-by-step instructions\n- Covers: Workspace setup, data connectors, analytics rules, workbooks, incidents, automation, hunting\n\nGreat prep for SC-200 and security operations roles.",
     "$5+", "Bootcamp"),

    ("w365-masterclass.jpg", "W365 Masterclass Assets",
     "Windows 365 Masterclass — Slides & Study Notes",
     "Complete presentation and study notes from the Windows 365 Masterclass.\n\nWhat you get:\n- Full PowerPoint presentation\n- Study notes companion\n- Covers: Cloud PC architecture, provisioning, management, troubleshooting, licensing\n\nCompanion to the W365 Masterclass video on YouTube.",
     "$5+", "Bootcamp"),

    # --- COURSE PRESENTATIONS ---
    ("sc900-course.jpg", "SC900 Course Training Presentation",
     "SC-900 Course — Full Training Presentation",
     "Complete training presentation for the SC-900: Microsoft Security, Compliance & Identity Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Security concepts, Entra ID, Defender, Purview, compliance\n\nCompanion to our free SC-900 YouTube full course.",
     "$5+", "Course"),

    ("pl900-course.jpg", "PL900 Course Training Presentation",
     "PL-900 Course — Full Training Presentation",
     "Complete training presentation for the PL-900: Microsoft Power Platform Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Power Apps, Power Automate, Power BI, Power Virtual Agents, Dataverse, AI Builder\n\nCompanion to our free PL-900 YouTube full course.",
     "$5+", "Course"),

    ("ms900-course.jpg", "MS900 Course Training Presentation",
     "MS-900 Course — Full Training Presentation",
     "Complete training presentation for the MS-900: Microsoft 365 Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Cloud concepts, M365 apps, security, compliance, licensing, support\n\nCompanion to our free MS-900 YouTube full course.",
     "$5+", "Course"),

    ("ms700-course.jpg", "MS700 Course Training Presentation",
     "MS-700 Course — Full Training Presentation",
     "Complete training presentation for the MS-700: Managing Microsoft Teams exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Teams lifecycle, calling, meetings, chat, compliance, governance\n\nCompanion to our free MS-700 YouTube full course.",
     "$5+", "Course"),

    ("ms500-course.jpg", "MS500 Course Training Presentation",
     "MS-500 Course — Full Training Presentation",
     "Complete training presentation for the MS-500: Microsoft 365 Security Administration exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Identity, threat protection, information protection, governance\n\nCompanion to our free MS-500 YouTube full course.",
     "$5+", "Course"),

    ("dp900-course.jpg", "DP900 Course Training Presentation",
     "DP-900 Course — Full Training Presentation",
     "Complete training presentation for the DP-900: Microsoft Azure Data Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Core data concepts, relational data, non-relational data, analytics\n\nCompanion to our free DP-900 YouTube full course.",
     "$5+", "Course"),

    ("az900-course.jpg", "AZ900 Course Training Presentation",
     "AZ-900 Course — Full Training Presentation",
     "Complete training presentation for the AZ-900: Microsoft Azure Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Cloud concepts, Azure architecture, compute, networking, storage, governance\n\nCompanion to our free AZ-900 YouTube full course.",
     "$5+", "Course"),

    ("az500-course.jpg", "AZ-500 Course Training Presentation",
     "AZ-500 Course — Full Training Presentation",
     "Complete training presentation for the AZ-500: Microsoft Azure Security Technologies exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Identity management, platform protection, security operations, data protection\n\nOur best-selling course presentation — used by 60+ people.",
     "$5+", "Course"),

    ("az400-course.jpg", "AZ-400 Course Training Presentation",
     "AZ-400 Course — Full Training Presentation",
     "Complete training presentation for the AZ-400: Microsoft DevOps Solutions exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Source control, CI/CD, dependency management, IaC, monitoring, feedback\n\nCompanion to our free AZ-400 YouTube full course.",
     "$5+", "Course"),

    ("az304-course.jpg", "AZ-304 Course Training Presentation",
     "AZ-304 Course — Azure Architect Design Presentation",
     "Training presentation for the AZ-304: Microsoft Azure Architect Design exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Monitoring, security, storage, data platforms, infrastructure, business continuity\n\nALL-TIME BEST SELLER — 461+ copies sold!\n\nNote: AZ-304 has been retired and replaced by AZ-305. This material is still valuable for learning Azure architecture concepts.",
     "$5+", "Course"),

    ("az303-course.jpg", "AZ-303 Course Training Presentation",
     "AZ-303 Course — Azure Architect Technologies Presentation",
     "Training presentation for the AZ-303: Microsoft Azure Architect Technologies exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: VMs, containers, web apps, storage, databases, identity, networking, migration\n\n#2 best seller — 340+ copies sold!\n\nNote: AZ-303 has been retired and replaced by AZ-305. Still excellent for Azure architecture learning.",
     "$5+", "Course"),

    ("az204-course.jpg", "AZ-204 Course Training Presentation",
     "AZ-204 Course — Full Training Presentation",
     "Complete training presentation for the AZ-204: Developing Solutions for Microsoft Azure exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Compute solutions, storage, security, monitoring, third-party integrations\n\nCompanion to our free AZ-204 YouTube full course.",
     "$5+", "Course"),

    ("az104-course.jpg", "AZ-104 Course Training Presentation",
     "AZ-104 Course — Full Training Presentation",
     "Complete training presentation for the AZ-104: Microsoft Azure Administrator exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: Identity, governance, compute, networking, storage, monitoring\n\nUsed by 138+ people. Companion to our free AZ-104 YouTube full course.",
     "$5+", "Course"),

    ("ai900-course.jpg", "AI-900 Course Training Presentation",
     "AI-900 Course — Full Training Presentation",
     "Complete training presentation for the AI-900: Microsoft Azure AI Fundamentals exam.\n\nWhat you get:\n- Full PowerPoint training deck\n- Covers all exam objectives\n- Topics: AI workloads, ML principles, computer vision, NLP, conversational AI, generative AI\n\nCompanion to our free AI-900 YouTube full course.",
     "$5+", "Course"),
]

cat_fills = {
    "Exam Q&A": "E6F7F5",
    "Interview": "FFF8E1",
    "Bootcamp": "E8F5E9",
    "Course": "E3F2FD",
}

row = 2
for i, (thumb, old_title, new_title, desc, price, cat) in enumerate(items, 1):
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=thumb).font = Font(size=10, color="006600")
    ws.cell(row=row, column=3, value=old_title).font = Font(size=10, color="888888")
    ws.cell(row=row, column=4, value=new_title).font = Font(size=10, bold=True)
    ws.cell(row=row, column=5, value=desc + FOOTER).alignment = wrap
    ws.cell(row=row, column=6, value=price)
    ws.cell(row=row, column=7, value=cat)
    if cat in cat_fills:
        ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=cat_fills[cat])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 200
    row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{row-1}"

out = r"C:\ssClawy\aguidetocloud-revamp\static\images\kofi\Ko-fi Product Updates.xlsx"
wb.save(out)
print(f"Created: {out}")
print(f"Rows: {row-2} items")
